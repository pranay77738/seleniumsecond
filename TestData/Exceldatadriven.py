import openpyxl

# load the data
book = openpyxl.load_workbook(
    "C:\\Users\\yjasw\\OneDrive\\Desktop\\Python\\workshet.xlsx"
)

# select the sheet from xl
sheet = book.active

# reading data
print(sheet.cell(row=1, column=1).value)

# write into xl
sheet.cell(row=5, column=6).value = "pranay"

# max rows, columns
print(sheet.max_row)
print(sheet.max_column)

# special case to read the value
print(sheet["B2"].value)
print("****************************************")

Dict = {}
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)
